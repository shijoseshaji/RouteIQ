# ============================================================
#  RouteIQ — Route Intelligence & Optimization Platform
#  First Mile Pickup Route Optimizer
#  Run: streamlit run routeiq_app.py
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
from math import radians, sin, cos, sqrt, atan2
import warnings
warnings.filterwarnings('ignore')
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import io

# ── PAGE CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="RouteIQ",
    page_icon="🗺️",
    layout="wide"
)

# ── CUSTOM CSS ───────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #F8F9FA; }
    .block-container { padding-top: 1rem; }
    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        text-align: center;
    }
    .stDownloadButton button {
        background-color: #1F4E79 !important;
        color: white !important;
        font-size: 16px !important;
        padding: 12px !important;
        border-radius: 8px !important;
        width: 100%;
    }
    .stButton > button[kind="primary"] {
        background-color: #1F4E79;
        color: white;
        font-size: 16px;
        padding: 12px;
        border-radius: 8px;
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)

# ── HEADER ───────────────────────────────────────────────────
st.markdown("""
<div style='text-align:center; padding: 10px 0 5px 0;'>
    <h1 style='color:#1F4E79; font-size:42px; margin-bottom:0;'>
        🗺️ RouteIQ
    </h1>
    <p style='color:#555; font-size:18px; margin-top:4px;'>
        Route Intelligence & Optimization Platform
    </p>
    <p style='color:#888; font-size:13px;'>
        First Mile Pickup Route Optimizer |
        Smart Executive Assignment |
        Cut-off Slot Planner
    </p>
</div>
<hr style='border:1px solid #DEE2E6;'>
""", unsafe_allow_html=True)


# ── SIDEBAR ───────────────────────────────────────────────────
with st.sidebar:
    st.image("https://img.icons8.com/color/96/route.png", width=60)
    st.markdown("## ⚙️ Configuration")
    st.markdown("---")

    st.markdown("**🕐 Shift Settings**")
    SHIFT_START_HOUR = st.number_input(
        "Shift Start Hour (24h)", min_value=6,  max_value=12, value=10)
    MAX_WORK_HOURS   = st.number_input(
        "Max Working Hours",      min_value=6,  max_value=12, value=9)

    st.markdown("---")
    st.markdown("**🚗 Route Constraints**")
    MAX_ROUTE_KM     = st.number_input(
        "Max Route KM / Executive", min_value=20, max_value=500, value=100)
    AVG_SPEED_KMPH   = st.number_input(
        "Avg Travel Speed (km/h)",  min_value=5,  max_value=80,  value=20)

    st.markdown("---")
    st.markdown("**📦 Productivity Settings**")
    MIN_PRODUCTIVITY = st.number_input(
        "Min Shipments / Executive", min_value=100, max_value=2000, value=450)
    PICKUP_TIME_SEC  = st.number_input(
        "Pickup Time / Shipment (sec)", min_value=5, max_value=120, value=10)

    st.markdown("---")
    st.markdown("**📍 Geo Settings**")
    OUTLIER_RADIUS_KM = st.number_input(
        "Outlier Radius from Hub (km)", min_value=10, max_value=300, value=60)

    st.markdown("---")
    st.markdown("""
    <div style='background:#EFF6FF; padding:12px;
                border-radius:8px; font-size:12px;'>
    <b>📋 Required Excel Format</b><br><br>
    <b>Sheet: Sellers</b><br>
    • seller_id<br>
    • volume<br>
    • latitude<br>
    • longitude<br>
    • hub_id<br><br>
    <b>Sheet: Hubs</b><br>
    • hub_id<br>
    • hub_lat<br>
    • hub_lon
    </div>
    """, unsafe_allow_html=True)

# Derived constants
SHIFT_END_SEC = (SHIFT_START_HOUR + MAX_WORK_HOURS) * 3600
SLOT_WINDOWS  = {
    12: (12, 14),
    14: (14, 16),
    16: (16, 18),
    18: (18, min(19, SHIFT_START_HOUR + MAX_WORK_HOURS)),
}
CUTOFF_SLOTS = [12, 14, 16, 18]
SLOT_LABELS  = {12: '12PM', 14: '2PM', 16: '4PM', 18: '6PM'}
SLOT_COLORS_HEX = {
    '12PM': '#D9F0A3',
    '2PM' : '#FFFACD',
    '4PM' : '#FFD9B3',
    '6PM' : '#FFB3B3'
}


# ── CORE FUNCTIONS ────────────────────────────────────────────
def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    a = sin((lat2-lat1)/2)**2 + cos(lat1)*cos(lat2)*sin((lon2-lon1)/2)**2
    return R * 2 * atan2(sqrt(a), sqrt(1-a))

def travel_seconds(lat1, lon1, lat2, lon2):
    return (haversine_km(lat1, lon1, lat2, lon2) / AVG_SPEED_KMPH) * 3600

def build_dist_matrix(hub_row, sellers):
    pts = [(hub_row.hub_lat, hub_row.hub_lon)] + \
          list(zip(sellers.latitude, sellers.longitude))
    n = len(pts)
    D = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            if i != j:
                D[i][j] = travel_seconds(
                    pts[i][0], pts[i][1], pts[j][0], pts[j][1])
    return D

def compute_route_km(route, hub_row, sellers_reset):
    if not route:
        return 0.0
    pts = [(hub_row.hub_lat, hub_row.hub_lon)] + \
          list(zip(sellers_reset.latitude, sellers_reset.longitude))
    km = haversine_km(pts[0][0], pts[0][1],
                      pts[route[0]][0], pts[route[0]][1])
    for k in range(len(route) - 1):
        km += haversine_km(pts[route[k]][0],  pts[route[k]][1],
                           pts[route[k+1]][0], pts[route[k+1]][1])
    km += haversine_km(pts[route[-1]][0], pts[route[-1]][1],
                       pts[0][0], pts[0][1])
    return round(km, 2)

def compute_route_hours(route, hub_row, sellers_reset):
    if not route:
        return 0.0
    D   = build_dist_matrix(hub_row, sellers_reset)
    t   = SHIFT_START_HOUR * 3600
    pos = 0
    for idx in route:
        t  += D[pos][idx]
        t  += sellers_reset.iloc[idx-1].volume * PICKUP_TIME_SEC
        pos = idx
    t += D[pos][0]
    return round((t - SHIFT_START_HOUR * 3600) / 3600, 2)

def secs_to_hhmm(secs):
    h = int(secs // 3600) % 24
    m = int((secs % 3600) // 60)
    return f"{h:02d}:{m:02d}"

def nn_route(available, D, start=0):
    route, remaining, cur = [], list(available), start
    while remaining:
        nxt = min(remaining, key=lambda j: D[cur][j])
        route.append(nxt)
        remaining.remove(nxt)
        cur = nxt
    return route

def route_dist_secs(route, D):
    if not route:
        return 0
    cost = D[0][route[0]]
    for k in range(len(route) - 1):
        cost += D[route[k]][route[k+1]]
    return cost + D[route[-1]][0]

def two_opt(route, D):
    if len(route) < 4:
        return route
    improved = True
    while improved:
        improved = False
        for i in range(1, len(route) - 1):
            for j in range(i + 1, len(route)):
                new = route[:i] + route[i:j+1][::-1] + route[j+1:]
                if route_dist_secs(new, D) < route_dist_secs(route, D):
                    route, improved = new, True
    return route

def filter_outliers(hub_row, sellers):
    sellers = sellers.copy()
    sellers['dist_km'] = sellers.apply(
        lambda r: haversine_km(hub_row.hub_lat, hub_row.hub_lon,
                               r.latitude, r.longitude), axis=1)
    valid    = sellers[sellers.dist_km <= OUTLIER_RADIUS_KM].drop(
        columns='dist_km')
    outliers = sellers[sellers.dist_km > OUTLIER_RADIUS_KM]
    return valid.reset_index(drop=True), outliers

def check_constraints(hub_row, sellers):
    sellers = sellers.reset_index(drop=True)
    D       = build_dist_matrix(hub_row, sellers)
    route   = nn_route(list(range(1, len(sellers)+1)), D)
    if len(route) > 3:
        route = two_opt(route, D)
    km  = compute_route_km(route, hub_row, sellers)
    hrs = compute_route_hours(route, hub_row, sellers)
    return km, hrs, (km > MAX_ROUTE_KM or hrs > MAX_WORK_HOURS)

def equal_volume_split(sellers):
    sellers  = sellers.reset_index(drop=True).copy()
    total_v  = sellers.volume.sum()
    target   = total_v / 2
    sorted_s = sellers.sort_values(
        ['longitude', 'latitude']).reset_index(drop=True)
    cum_vol, half_a_idx, half_b_idx = 0, [], []
    for _, row in sorted_s.iterrows():
        if cum_vol < target:
            half_a_idx.append(row.seller_id)
            cum_vol += row.volume
        else:
            half_b_idx.append(row.seller_id)
    half_a = sellers[
        sellers.seller_id.isin(half_a_idx)].reset_index(drop=True)
    half_b = sellers[
        sellers.seller_id.isin(half_b_idx)].reset_index(drop=True)
    return half_a, half_b

def recursive_split(hub_row, sellers, base_exec_id, split_log,
                    depth=0, max_depth=6):
    if depth > max_depth:
        return [(base_exec_id, sellers, depth > 0)]
    km, hrs, breached = check_constraints(hub_row, sellers)
    if not breached:
        return [(base_exec_id, sellers, depth > 0)]
    reason = []
    if km  > MAX_ROUTE_KM:   reason.append(f"{km:.1f}km>{MAX_ROUTE_KM}km")
    if hrs > MAX_WORK_HOURS: reason.append(f"{hrs:.1f}h>{MAX_WORK_HOURS}h")
    half_a, half_b = equal_volume_split(sellers)
    id_a = f"{base_exec_id}_S{depth+1}A"
    id_b = f"{base_exec_id}_S{depth+1}B"
    split_log.append({
        'parent_exec'     : base_exec_id,
        'child_a'         : id_a,
        'child_b'         : id_b,
        'reason'          : ' & '.join(reason),
        'parent_sellers'  : len(sellers),
        'parent_volume'   : sellers.volume.sum(),
        'child_a_sellers' : len(half_a),
        'child_a_volume'  : half_a.volume.sum(),
        'child_b_sellers' : len(half_b),
        'child_b_volume'  : half_b.volume.sum(),
    })
    results  = []
    results += recursive_split(
        hub_row, half_a, id_a, split_log, depth+1)
    results += recursive_split(
        hub_row, half_b, id_b, split_log, depth+1)
    return results

def assign_slots_to_executive(hub_row, exec_sellers, exec_id, was_split):
    exec_sellers = exec_sellers.reset_index(drop=True).copy()
    D            = build_dist_matrix(hub_row, exec_sellers)
    n            = len(exec_sellers)
    global_route = nn_route(list(range(1, n+1)), D)
    if len(global_route) > 3:
        global_route = two_opt(global_route, D)
    route_km  = compute_route_km(global_route, hub_row, exec_sellers)
    route_hrs = compute_route_hours(global_route, hub_row, exec_sellers)

    slot_assignments = {}
    remaining_route  = list(global_route)

    for slot_hour in CUTOFF_SLOTS:
        _, cutoff_h      = SLOT_WINDOWS[slot_hour]
        effective_cutoff = min(cutoff_h * 3600, SHIFT_END_SEC)
        current_time     = (SHIFT_START_HOUR * 3600
                            if slot_hour == CUTOFF_SLOTS[0]
                            else SLOT_WINDOWS[slot_hour][0] * 3600)
        current_pos   = 0
        still_pending = []
        for dist_idx in remaining_route:
            tt  = D[current_pos][dist_idx]
            pd_ = exec_sellers.iloc[dist_idx-1].volume * PICKUP_TIME_SEC
            tb  = D[dist_idx][0]
            if current_time + tt + pd_ + tb <= effective_cutoff:
                slot_assignments[dist_idx] = slot_hour
                current_time += tt + pd_
                current_pos   = dist_idx
            else:
                still_pending.append(dist_idx)
        remaining_route = still_pending

    for dist_idx in remaining_route:
        slot_assignments[dist_idx] = 18

    def slot_vol(slot):
        return sum(exec_sellers.iloc[i-1].volume
                   for i, s in slot_assignments.items() if s == slot)

    for target_slot in [12, 14, 16]:
        _, cutoff_h      = SLOT_WINDOWS[target_slot]
        effective_cutoff = min(cutoff_h * 3600, SHIFT_END_SEC)
        t_start          = (SHIFT_START_HOUR * 3600
                            if target_slot == CUTOFF_SLOTS[0]
                            else SLOT_WINDOWS[target_slot][0] * 3600)
        slot_in_order = [i for i in global_route
                         if slot_assignments.get(i) == target_slot]
        used_time, used_pos = t_start, 0
        for i in slot_in_order:
            used_time += (D[used_pos][i] +
                          exec_sellers.iloc[i-1].volume * PICKUP_TIME_SEC)
            used_pos   = i
        for dist_idx in [i for i in global_route
                         if slot_assignments.get(i) == 18]:
            tt  = D[used_pos][dist_idx]
            pd_ = exec_sellers.iloc[dist_idx-1].volume * PICKUP_TIME_SEC
            tb  = D[dist_idx][0]
            if used_time + tt + pd_ + tb <= effective_cutoff:
                slot_assignments[dist_idx] = target_slot
                used_time += tt + pd_
                used_pos   = dist_idx

    results = []
    for slot_hour in CUTOFF_SLOTS:
        sellers_in_slot = [i for i in global_route
                           if slot_assignments.get(i) == slot_hour]
        if len(sellers_in_slot) > 3:
            sellers_in_slot = two_opt(sellers_in_slot, D)
        current_time = (SHIFT_START_HOUR * 3600
                        if slot_hour == CUTOFF_SLOTS[0]
                        else SLOT_WINDOWS[slot_hour][0] * 3600)
        current_pos  = 0
        for seq, dist_idx in enumerate(sellers_in_slot, 1):
            s          = exec_sellers.iloc[dist_idx - 1]
            from_lat   = (hub_row.hub_lat if current_pos == 0
                          else exec_sellers.iloc[current_pos-1].latitude)
            from_lon   = (hub_row.hub_lon if current_pos == 0
                          else exec_sellers.iloc[current_pos-1].longitude)
            from_label = ('Hub' if current_pos == 0
                          else exec_sellers.iloc[current_pos-1].seller_id)
            transit_km   = haversine_km(
                from_lat, from_lon, s.latitude, s.longitude)
            transit_secs = travel_seconds(
                from_lat, from_lon, s.latitude, s.longitude)
            arrival_time = current_time + transit_secs
            pickup_end   = arrival_time + s.volume * PICKUP_TIME_SEC
            results.append({
                'seller_id'           : s.seller_id,
                'hub_id'              : hub_row.hub_id,
                'exec_id'             : exec_id,
                'forced_split'        : 'YES' if was_split else 'NO',
                'volume'              : s.volume,
                'latitude'            : s.latitude,
                'longitude'           : s.longitude,
                'slot_sequence'       : seq,
                'assigned_cutoff'     : SLOT_LABELS[slot_hour],
                'cutoff_hour'         : slot_hour,
                'from_point'          : from_label,
                'transit_distance_km' : round(transit_km, 2),
                'transit_time_mins'   : round(transit_secs / 60, 1),
                'estimated_arrival'   : secs_to_hhmm(arrival_time),
                'pickup_duration_mins': round(
                    s.volume * PICKUP_TIME_SEC / 60, 1),
                'estimated_pickup_end': secs_to_hhmm(pickup_end),
                'total_route_km'      : route_km,
                'total_route_hrs'     : route_hrs,
            })
            current_time = pickup_end
            current_pos  = dist_idx
    return results

def cluster_sellers_intact(hub_row, sellers, hub_id):
    sellers   = sellers.reset_index(drop=True).copy()
    total_vol = sellers.volume.sum()
    n_exec    = max(1, int(np.floor(total_vol / MIN_PRODUCTIVITY)))
    if n_exec == 1:
        sellers['cluster'] = 0
    else:
        coords = sellers[['latitude', 'longitude']].values
        km_cl  = KMeans(n_clusters=n_exec, random_state=42, n_init=10)
        sellers['cluster'] = km_cl.fit_predict(coords)
    while True:
        cluster_vols = sellers.groupby('cluster')['volume'].sum()
        under = cluster_vols[
            cluster_vols < MIN_PRODUCTIVITY].index.tolist()
        if not under:
            break
        centroids = sellers.groupby('cluster')[
            ['latitude', 'longitude']].mean()
        worst  = cluster_vols[under].idxmin()
        others = [c for c in centroids.index if c != worst]
        if not others:
            break
        wc      = centroids.loc[worst]
        nearest = min(others, key=lambda c: haversine_km(
            wc.latitude, wc.longitude,
            centroids.loc[c].latitude, centroids.loc[c].longitude))
        sellers.loc[sellers.cluster == worst, 'cluster'] = nearest
    unique_clusters = sorted(sellers.cluster.unique())
    cmap = {c: f"{hub_id}_EXE_{i+1}"
            for i, c in enumerate(unique_clusters)}
    sellers['exec_id'] = sellers['cluster'].map(cmap)
    return sellers.drop(columns='cluster')


# ── EXCEL BUILDER ─────────────────────────────────────────────
def build_excel(output_df, productivity, slot_summary,
                sp_cnt, sp_vol, split_log, all_outliers,
                route_summary):
    HDR_FILL = PatternFill('solid', start_color='1F4E79')
    HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    CTR      = Alignment(horizontal='center', vertical='center')
    SLOT_CLR = {'12PM':'D9F0A3','2PM':'FFFACD',
                '4PM':'FFD9B3','6PM':'FFB3B3'}
    WARN_CLR, OK_CLR, SPL_CLR = 'FFC7CE', 'C6EFCE', 'FFE0B2'

    def sh(ws, labels, widths):
        for c, lbl in enumerate(labels, 1):
            cell = ws.cell(row=1, column=c, value=lbl)
            cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, CTR
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

    def fr(ws, r, vals, slot=None, bg=None):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = CTR
        clr = SLOT_CLR.get(slot, bg) if slot else bg
        if clr:
            for c in range(1, len(vals)+1):
                ws.cell(row=r, column=c).fill = PatternFill(
                    'solid', start_color=clr)

    wb  = Workbook()
    sdf = output_df.sort_values(
        ['hub_id','exec_id','cutoff_hour','slot_sequence'])

    # Sheet 1: Seller Assignments
    ws1 = wb.active
    ws1.title = 'Seller Assignments'
    sh(ws1, [
        'Seller ID','Hub','Executive','Forced Split?','Volume',
        'Latitude','Longitude','Route Seq.','Slot Seq.',
        'Assigned Cut-off','From Point',
        'Transit Dist (km)','Transit Time (mins)',
        'Est. Arrival','Pickup Dur (mins)','Est. Pickup End',
        'Total Route KM','Total Route Hrs'
    ], [18,22,28,12,8,12,12,10,10,16,20,18,20,14,18,16,15,15])
    for r, row in enumerate(sdf.itertuples(index=False), 2):
        fr(ws1, r, [
            row.seller_id, row.hub_id, row.exec_id,
            row.forced_split, row.volume,
            row.latitude, row.longitude,
            row.route_sequence, row.slot_sequence,
            row.assigned_cutoff, row.from_point,
            row.transit_distance_km, row.transit_time_mins,
            row.estimated_arrival, row.pickup_duration_mins,
            row.estimated_pickup_end,
            row.total_route_km, row.total_route_hrs
        ], slot=row.assigned_cutoff)

    # Sheet 2: Executive Summary
    ws2 = wb.create_sheet('Executive Summary')
    sh(ws2, [
        'Executive','Hub','Forced Split?',
        'Total Sellers','Total Volume',
        'Total Route KM','KM OK?',
        'Total Route Hrs','Hours OK?',
        'Total Travel (mins)',
        '12PM sellers','12PM vol',
        '2PM sellers', '2PM vol',
        '4PM sellers', '4PM vol',
        '6PM sellers', '6PM vol',
        'Meets Min?'
    ], [28,22,12,14,14,15,10,15,10,18,
        13,10,13,10,13,10,13,10,12])
    for r, row in enumerate(productivity.itertuples(index=False), 2):
        eid = row.exec_id
        sc = lambda s: int(sp_cnt.loc[eid,s]) if eid in sp_cnt.index else 0
        sv = lambda s: int(sp_vol.loc[eid,s]) if eid in sp_vol.index else 0
        any_issue = not row.km_ok or not row.hours_ok
        bg = WARN_CLR if any_issue else (
             SPL_CLR if row.forced_split=='YES' else OK_CLR)
        fr(ws2, r, [
            eid, row.hub_id, row.forced_split,
            row.total_sellers, row.total_volume,
            row.total_route_km, 'YES' if row.km_ok   else 'NO ⚠️',
            row.total_route_hrs,'YES' if row.hours_ok else 'NO ⚠️',
            row.total_travel_mins,
            sc('12PM'),sv('12PM'),sc('2PM'),sv('2PM'),
            sc('4PM'), sv('4PM'), sc('6PM'),sv('6PM'),
            'YES' if row.meets_min else 'NO ⚠️'
        ], bg=bg)

    # Sheet 3: Slot Distribution
    ws3 = wb.create_sheet('Slot Distribution')
    sh(ws3, ['Hub','Executive','Cut-off Slot',
             'Sellers','Volume',
             'Total Distance (km)','Total Travel (mins)'],
       [22,28,14,10,10,18,18])
    for r, row in enumerate(slot_summary.itertuples(index=False), 2):
        fr(ws3, r, [
            row.hub_id, row.exec_id, row.assigned_cutoff,
            row.sellers, row.volume,
            row.total_dist_km, row.total_travel_mins
        ], slot=row.assigned_cutoff)

    # Sheet 4: Transit Detail
    ws4 = wb.create_sheet('Transit Detail')
    sh(ws4, [
        'Executive','Hub','Slot','Slot Seq.',
        'From Point','To Seller',
        'Transit Dist (km)','Transit Time (mins)',
        'Est. Arrival','Volume',
        'Pickup Dur (mins)','Est. Pickup End'
    ], [28,22,10,10,20,20,18,20,14,8,18,16])
    for r, row in enumerate(sdf.itertuples(index=False), 2):
        fr(ws4, r, [
            row.exec_id, row.hub_id,
            row.assigned_cutoff, row.slot_sequence,
            row.from_point, row.seller_id,
            row.transit_distance_km, row.transit_time_mins,
            row.estimated_arrival, row.volume,
            row.pickup_duration_mins, row.estimated_pickup_end
        ], slot=row.assigned_cutoff)

    # Sheet 5: Route Summary
    ws5 = wb.create_sheet('Route Summary')
    sh(ws5, [
        'Executive','Hub','Forced Split?',
        'Total Stops','Total Volume',
        'Total Route KM','KM OK?',
        'Total Route Hrs','Hours OK?',
        'Total Travel (mins)','Total Pickup (mins)',
        'Total Day (mins)','Shift Start','Last Pickup End'
    ], [28,22,12,12,14,15,10,15,10,20,20,16,12,16])
    for r, row in enumerate(route_summary.itertuples(index=False), 2):
        any_issue = not row.km_ok or not row.hours_ok
        bg = WARN_CLR if any_issue else (
             SPL_CLR if row.forced_split=='YES' else OK_CLR)
        fr(ws5, r, [
            row.exec_id, row.hub_id, row.forced_split,
            row.total_stops, row.total_volume,
            row.total_route_km,  'YES' if row.km_ok    else 'NO ⚠️',
            row.total_route_hrs, 'YES' if row.hours_ok else 'NO ⚠️',
            row.total_travel_mins, row.total_pickup_mins,
            row.total_day_mins, row.shift_start,
            row.last_pickup_end
        ], bg=bg)

    # Sheet 6: Split Log
    if split_log:
        ws6 = wb.create_sheet('Split Log')
        sh(ws6, [
            'Parent Executive','Reason',
            'Child A','Child A Sellers','Child A Volume',
            'Child B','Child B Sellers','Child B Volume',
            'Parent Sellers','Parent Volume'
        ], [30,30,30,15,15,30,15,15,15,15])
        for r, entry in enumerate(split_log, 2):
            fr(ws6, r, [
                entry['parent_exec'], entry['reason'],
                entry['child_a'],     entry['child_a_sellers'],
                entry['child_a_volume'],
                entry['child_b'],     entry['child_b_sellers'],
                entry['child_b_volume'],
                entry['parent_sellers'], entry['parent_volume']
            ], bg=SPL_CLR)

    # Sheet 7: Outliers
    if all_outliers:
        outlier_df = pd.concat(all_outliers, ignore_index=True)
        ws7 = wb.create_sheet('Outliers - Manual Review')
        sh(ws7, [
            'Seller ID','Hub','Latitude','Longitude',
            'Volume','Dist from Hub (km)','Reason'
        ], [18,22,12,12,10,20,30])
        for r, row in enumerate(outlier_df.itertuples(index=False), 2):
            fr(ws7, r, [
                row.seller_id, row.hub_id,
                row.latitude, row.longitude,
                row.volume, round(row.dist_km, 1), row.reason
            ], bg='FFE0E0')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── MAIN APP BODY ─────────────────────────────────────────────
uploaded = st.file_uploader(
    "📂 Upload your Excel file",
    type=['xlsx'],
    help="File must have two sheets: 'Sellers' and 'Hubs'"
)

if not uploaded:
    # Landing page
    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    cards = [
        ("🗺️", "Smart Routing",
         "Nearest-neighbour + 2-opt optimisation for shortest path"),
        ("⏱️", "Slot Planning",
         "Fills 12PM → 2PM → 4PM slots first. 6PM gets least volume"),
        ("📏", "Constraint Engine",
         "Auto-splits if route exceeds 100km or 9 working hours"),
        ("📊", "Full Reporting",
         "7-sheet Excel output with transit times, distances & maps"),
    ]
    for col, (icon, title, desc) in zip([c1,c2,c3,c4], cards):
        col.markdown(f"""
        <div style='background:white; border-radius:12px;
                    padding:20px; text-align:center;
                    box-shadow:0 2px 8px rgba(0,0,0,0.08);
                    height:160px;'>
            <div style='font-size:32px;'>{icon}</div>
            <div style='font-weight:bold; color:#1F4E79;
                        margin:8px 0 4px;'>{title}</div>
            <div style='font-size:12px; color:#666;'>{desc}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("👆 Upload your Excel file above to get started. "
            "Adjust settings in the left sidebar before running.")

else:
    try:
        sellers_df = pd.read_excel(uploaded, sheet_name='Sellers')
        uploaded.seek(0)
        hubs_df    = pd.read_excel(uploaded, sheet_name='Hubs')
        sellers_df = sellers_df.drop_duplicates(
            subset='seller_id').reset_index(drop=True)

        # Data preview
        st.markdown("### 📋 Data Preview")
        pc1, pc2 = st.columns([3, 1])
        with pc1:
            st.dataframe(sellers_df.head(8), use_container_width=True)
        with pc2:
            st.dataframe(hubs_df, use_container_width=True)

        # Summary metrics
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Sellers",   f"{len(sellers_df):,}")
        m2.metric("Total Volume",    f"{sellers_df.volume.sum():,}")
        m3.metric("Total Hubs",      sellers_df.hub_id.nunique())
        m4.metric("Avg Vol/Seller",
                  f"{sellers_df.volume.mean():.1f}")

        st.markdown("---")

        run = st.button(
            "🚀 Run RouteIQ Optimization",
            type="primary",
            use_container_width=True
        )

        if run:
            all_results  = []
            all_outliers = []
            split_log    = []

            prog    = st.progress(0)
            status  = st.status(
                "🔄 RouteIQ is optimizing routes...",
                expanded=True)
            total_h = len(hubs_df)

            for hi, (_, hub_row) in enumerate(hubs_df.iterrows()):
                hub_sellers = sellers_df[
                    sellers_df.hub_id == hub_row.hub_id].copy()
                if hub_sellers.empty:
                    continue

                status.write(
                    f"🏠 Hub **{hub_row.hub_id}** "
                    f"— {len(hub_sellers)} sellers")

                hub_sellers, outliers = filter_outliers(
                    hub_row, hub_sellers)
                if not outliers.empty:
                    o = outliers.copy()
                    o['hub_id'] = hub_row.hub_id
                    o['reason'] = f'>{OUTLIER_RADIUS_KM}km from hub'
                    all_outliers.append(o)

                if hub_sellers.empty:
                    continue

                hub_sellers = cluster_sellers_intact(
                    hub_row, hub_sellers, hub_row.hub_id)

                final_execs = []
                for exec_id, exec_s in hub_sellers.groupby('exec_id'):
                    exec_s      = exec_s.reset_index(drop=True)
                    es_hub      = exec_s.copy()
                    es_hub['hub_id'] = hub_row.hub_id
                    res = recursive_split(
                        hub_row, es_hub, exec_id, split_log)
                    for rid, rs, rsplit in res:
                        rs = rs.drop(
                            columns=['hub_id'],
                            errors='ignore').reset_index(drop=True)
                        final_execs.append((rid, rs, rsplit))

                for exec_id, exec_s, was_split in final_execs:
                    status.write(
                        f"  ✓ {exec_id} | "
                        f"{len(exec_s)} sellers | "
                        f"vol={exec_s.volume.sum()}"
                        f"{' [SPLIT]' if was_split else ''}")
                    exec_results = assign_slots_to_executive(
                        hub_row, exec_s, exec_id, was_split)
                    all_results.extend(exec_results)

                prog.progress((hi + 1) / total_h)

            status.update(
                label="✅ RouteIQ optimization complete!",
                state="complete")

            # Build output dataframe
            output_df = pd.DataFrame(all_results)
            output_df = output_df.sort_values(
                ['hub_id','exec_id','cutoff_hour','slot_sequence'])
            output_df['route_sequence'] = \
                output_df.groupby('exec_id').cumcount() + 1

            # Productivity
            productivity = output_df.groupby('exec_id').agg(
                hub_id            =('hub_id',            'first'),
                forced_split      =('forced_split',       'first'),
                total_sellers     =('seller_id',          'count'),
                total_volume      =('volume',             'sum'),
                total_distance_km =('transit_distance_km','sum'),
                total_travel_mins =('transit_time_mins',  'sum'),
                total_route_km    =('total_route_km',     'first'),
                total_route_hrs   =('total_route_hrs',    'first'),
            ).reset_index()
            productivity['km_ok']    = \
                productivity.total_route_km  <= MAX_ROUTE_KM
            productivity['hours_ok'] = \
                productivity.total_route_hrs <= MAX_WORK_HOURS
            productivity['meets_min'] = (
                (productivity.total_volume >= MIN_PRODUCTIVITY) |
                (productivity.forced_split == 'YES'))

            slot_summary = output_df.groupby(
                ['hub_id','exec_id','assigned_cutoff']).agg(
                sellers           =('seller_id',          'count'),
                volume            =('volume',             'sum'),
                total_dist_km     =('transit_distance_km','sum'),
                total_travel_mins =('transit_time_mins',  'sum'),
            ).reset_index()

            sp_cnt = output_df.groupby(
                ['exec_id','assigned_cutoff'])[
                'seller_id'].count().unstack(fill_value=0)
            sp_vol = output_df.groupby(
                ['exec_id','assigned_cutoff'])[
                'volume'].sum().unstack(fill_value=0)
            for slot in ['12PM','2PM','4PM','6PM']:
                if slot not in sp_cnt.columns: sp_cnt[slot] = 0
                if slot not in sp_vol.columns: sp_vol[slot] = 0

            route_summary = output_df.groupby('exec_id').agg(
                hub_id            =('hub_id',            'first'),
                forced_split      =('forced_split',       'first'),
                total_stops       =('seller_id',          'count'),
                total_volume      =('volume',             'sum'),
                total_route_km    =('total_route_km',     'first'),
                total_route_hrs   =('total_route_hrs',    'first'),
                total_travel_mins =('transit_time_mins',  'sum'),
                total_pickup_mins =('pickup_duration_mins','sum'),
                last_pickup_end   =('estimated_pickup_end','last'),
            ).reset_index()
            route_summary['total_day_mins'] = (
                route_summary['total_travel_mins'] +
                route_summary['total_pickup_mins']).round(1)
            route_summary['shift_start'] = \
                f"{SHIFT_START_HOUR:02d}:00"
            route_summary['km_ok']    = \
                route_summary.total_route_km  <= MAX_ROUTE_KM
            route_summary['hours_ok'] = \
                route_summary.total_route_hrs <= MAX_WORK_HOURS

            # ── KPI Banner ───────────────────────────────────
            st.markdown("---")
            st.markdown("### 📊 Optimization Results")

            k1,k2,k3,k4,k5,k6 = st.columns(6)
            k1.metric("Sellers Processed", f"{len(output_df):,}")
            k2.metric("Total Volume",
                      f"{output_df.volume.sum():,}")
            k3.metric("Executives Assigned",
                      output_df.exec_id.nunique())
            k4.metric("Forced Splits",     len(split_log))
            k5.metric("Outliers Flagged",
                      sum(len(o) for o in all_outliers))
            k6.metric("Constraint Breaches",
                      int((~productivity.km_ok |
                           ~productivity.hours_ok).sum()))

            # ── Result Tabs ──────────────────────────────────
            t1,t2,t3,t4,t5 = st.tabs([
                "📋 Assignments",
                "👤 Executives",
                "📦 Slot Distribution",
                "📈 Charts",
                "🗺️ Map"
            ])

            with t1:
                st.dataframe(
                    output_df[[
                        'seller_id','hub_id','exec_id',
                        'forced_split','volume',
                        'assigned_cutoff','route_sequence',
                        'slot_sequence','from_point',
                        'transit_distance_km','transit_time_mins',
                        'estimated_arrival','estimated_pickup_end',
                        'total_route_km','total_route_hrs'
                    ]],
                    use_container_width=True,
                    height=450
                )

            with t2:
                def colour_exec(row):
                    if not row.km_ok or not row.hours_ok:
                        return ['background-color:#FFC7CE']*len(row)
                    elif row.forced_split == 'YES':
                        return ['background-color:#FFE0B2']*len(row)
                    return ['background-color:#C6EFCE']*len(row)
                st.dataframe(
                    productivity.style.apply(colour_exec, axis=1),
                    use_container_width=True,
                    height=450
                )
                st.caption(
                    "🟢 Green = all OK  |  "
                    "🟠 Orange = forced split  |  "
                    "🔴 Red = constraint breach")

            with t3:
                st.dataframe(
                    slot_summary,
                    use_container_width=True,
                    height=450
                )

            with t4:
                execs       = route_summary.exec_id.tolist()
                x           = np.arange(len(execs))
                slot_order  = ['12PM','2PM','4PM','6PM']
                slot_colors = ['#7BC67E','#FFD966',
                               '#FF9F40','#FF6B6B']
                width       = 0.2

                fig, axes = plt.subplots(
                    1, 3, figsize=(18, 5),
                    facecolor='#F8F9FA')

                # Chart 1: Route KM
                c_km = ['#FF6B6B' if not ok else '#4472C4'
                        for ok in route_summary.km_ok]
                axes[0].bar(x, route_summary.total_route_km,
                            color=c_km, edgecolor='white',
                            linewidth=0.5)
                axes[0].axhline(
                    MAX_ROUTE_KM, color='red',
                    linestyle='--', linewidth=1.5,
                    label=f'{MAX_ROUTE_KM}km limit')
                axes[0].set_xticks(x)
                axes[0].set_xticklabels(
                    execs, rotation=35, ha='right', fontsize=7)
                axes[0].set_title(
                    'Route KM per Executive',
                    fontweight='bold', pad=10)
                axes[0].set_ylabel('KM')
                axes[0].legend(fontsize=8)
                axes[0].grid(axis='y', alpha=0.3)
                axes[0].set_facecolor('#FAFAFA')

                # Chart 2: Working Hours
                c_hrs = ['#FF6B6B' if not ok else '#70AD47'
                         for ok in route_summary.hours_ok]
                axes[1].bar(x, route_summary.total_route_hrs,
                            color=c_hrs, edgecolor='white',
                            linewidth=0.5)
                axes[1].axhline(
                    MAX_WORK_HOURS, color='red',
                    linestyle='--', linewidth=1.5,
                    label=f'{MAX_WORK_HOURS}h limit')
                axes[1].set_xticks(x)
                axes[1].set_xticklabels(
                    execs, rotation=35, ha='right', fontsize=7)
                axes[1].set_title(
                    'Working Hours per Executive',
                    fontweight='bold', pad=10)
                axes[1].set_ylabel('Hours')
                axes[1].legend(fontsize=8)
                axes[1].grid(axis='y', alpha=0.3)
                axes[1].set_facecolor('#FAFAFA')

                # Chart 3: Slot Volume
                for i, (slot, color) in enumerate(
                        zip(slot_order, slot_colors)):
                    vals = [int(sp_vol.loc[e, slot])
                            if e in sp_vol.index else 0
                            for e in execs]
                    axes[2].bar(
                        x + i*width, vals, width,
                        label=slot, color=color,
                        edgecolor='white', linewidth=0.5)
                axes[2].set_xticks(x + width*1.5)
                axes[2].set_xticklabels(
                    execs, rotation=35, ha='right', fontsize=7)
                axes[2].set_title(
                    'Volume per Slot per Executive',
                    fontweight='bold', pad=10)
                axes[2].set_ylabel('Shipments')
                axes[2].legend(fontsize=8)
                axes[2].grid(axis='y', alpha=0.3)
                axes[2].set_facecolor('#FAFAFA')

                plt.suptitle(
                    'RouteIQ — Executive Constraint Dashboard',
                    fontsize=13, fontweight='bold',
                    color='#1F4E79', y=1.02)
                plt.tight_layout()
                st.pyplot(fig)
                plt.close()

            with t5:
                slot_order  = ['12PM','2PM','4PM','6PM']
                colors_map  = {
                    '12PM':'#7BC67E','2PM':'#FFD966',
                    '4PM':'#FF9F40', '6PM':'#FF6B6B'
                }
                n_hubs = len(hubs_df)
                fig, axes_map = plt.subplots(
                    1, n_hubs,
                    figsize=(10*n_hubs, 7),
                    squeeze=False,
                    facecolor='#F8F9FA')

                for col, (_, hub_row) in enumerate(
                        hubs_df.iterrows()):
                    ax       = axes_map[0][col]
                    hub_data = output_df[
                        output_df.hub_id == hub_row.hub_id]

                    for slot, grp in hub_data.groupby(
                            'assigned_cutoff'):
                        ax.scatter(
                            grp.longitude, grp.latitude,
                            c=colors_map.get(slot,'#CCCCCC'),
                            label=slot, s=60,
                            edgecolors='white',
                            linewidths=0.4, zorder=3)

                    ax.scatter(
                        hub_row.hub_lon, hub_row.hub_lat,
                        marker='*', s=500, c='#1F4E79',
                        zorder=5, label='Hub')
                    ax.annotate(
                        hub_row.hub_id,
                        (hub_row.hub_lon, hub_row.hub_lat),
                        textcoords='offset points',
                        xytext=(8, 8), fontsize=9,
                        fontweight='bold', color='#1F4E79')

                    counts = hub_data.groupby(
                        'assigned_cutoff')[
                        'seller_id'].count().to_dict()
                    vols   = hub_data.groupby(
                        'assigned_cutoff')[
                        'volume'].sum().to_dict()
                    lines  = [
                        f"{s}: {counts.get(s,0)} sellers "
                        f"/ {vols.get(s,0)} vol"
                        for s in slot_order
                        if counts.get(s, 0) > 0
                    ]
                    ax.set_title(
                        hub_row.hub_id + '\n' +
                        '  |  '.join(lines),
                        fontsize=9, fontweight='bold',
                        color='#1F4E79')
                    ax.set_xlabel('Longitude', fontsize=9)
                    ax.set_ylabel('Latitude',  fontsize=9)
                    ax.legend(loc='upper left', fontsize=8)
                    ax.grid(alpha=0.3)
                    ax.set_facecolor('#FAFAFA')

                plt.tight_layout()
                st.pyplot(fig)
                plt.close()

            # ── Download ─────────────────────────────────────
            st.markdown("---")
            excel_buf = build_excel(
                output_df, productivity, slot_summary,
                sp_cnt, sp_vol, split_log, all_outliers,
                route_summary)

            st.download_button(
                label="⬇️  Download Full RouteIQ Report (Excel)",
                data=excel_buf,
                file_name="RouteIQ_optimized_routes.xlsx",
                mime=("application/vnd.openxmlformats-"
                      "officedocument.spreadsheetml.sheet"),
                use_container_width=True,
                type="primary"
            )

    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
        st.info(
            "Please ensure your Excel file has sheets named "
            "**Sellers** and **Hubs** with the correct columns.")
